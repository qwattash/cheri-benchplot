import re
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from matplotlib import colors as mcolors
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils import get_column_letter

from .backend import FigureManager, Mosaic, ViewRenderer
from .data_view import CellData, DataView, TableDataView


class ExcelFigureManager(FigureManager):
    """
    Draw plots to static HTML files.
    """

    # @dataclass
    # class DrawContext(Surface.DrawContext):
    #     writer: pd.ExcelWriter

    def __init__(self, config):
        super().__init__(config)
        self.writer = None

    def allocate_cells(self, mosaic: Mosaic):
        for subplot in mosaic:
            subplot.cell = ExcelCellData(title=subplot.get_cell_title())

    def draw(self, mosaic, title, dest):
        super().draw(mosaic, title, dest)
        if self.config.split_subplots:
            self.logger.warning("Split subplots unsupported for excel figures")
        with pd.ExcelWriter(dest.with_suffix(".xlsx"), mode="w", engine="openpyxl") as writer:
            for subplot in mosaic:
                subplot.cell.writer = writer
                subplot.cell.draw()


class ExcelCellData(CellData):
    """
    Base HTML dataset rendering class. Add wrapper functions to allow running the rendering
    step within the jinja templates, so that we can access cell and view properties within the
    template if needed.
    """
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._renderers = {"table": SpreadsheetTableRenderer}
        # Set at drawing time
        self.writer = None

    def draw(self):
        if len(self.views) > 1:
            self.surface.logger.warning("Only a single plot view is supported for each excel surface cell")
        try:
            view = self.views[0]
        except IndexError:
            return
        r = self.get_renderer(view)
        r.render(view, self)


class SpreadsheetTableRenderer(ViewRenderer):
    """
    Render table in a spreadsheet.
    """
    def _gen_fill_styles(self, view):
        """
        Generate a dataframe with the same index and columns as the view dataframe, containing the
        colors for each cell.
        """
        if view.legend_info is None:
            return None

        legend_df = view.legend_info.info_df
        fill_styles = []
        for color in legend_df["colors"]:
            # openpyxl uses aRGB colors while matplotlib uses RGBa
            # we need to swap the alpha here
            rgba_hex = [int(round(c * 255)) for c in color]
            argb_hex = rgba_hex[-1:] + rgba_hex[0:3]
            hexcolor = "".join(map(lambda c: f"{c:02x}", argb_hex))
            style = PatternFill(start_color=hexcolor, end_color=hexcolor, fill_type="solid")
            fill_styles.append(style)
        fs = pd.DataFrame({"fill_style": fill_styles}, index=legend_df.index)

        if view.legend_level:
            # Before joining we need to align the fs and view.df column indexes
            extra_levels = [['']] * (view.df.columns.nlevels - 1)
            col_idx = pd.MultiIndex.from_product([fs.columns] + extra_levels, names=view.df.columns.names)
            fs.columns = col_idx
            fill_df = view.df.join(fs, on=view.legend_level)
        elif "column" in fs.index.names:
            assert len(fs.index.names) == 1
            index = view.df.index.repeat(len(fs))
            fill_df = pd.DataFrame({"column": None}, index=index)
            fill_df["column"] = np.tile(fs.index, len(view.df))
            fill_df = fill_df.join(fs, on="column")
        else:
            raise ValueError("Malformed legend_info: missing legend_level or 'column' level")
        return fill_df

    def render(self, view: TableDataView, cell):
        """
        Render the dataframe as a table in an excel sheet.
        """
        assert isinstance(view, TableDataView), "Table renderer can only handle TableDataView"
        excel_writer = cell.writer

        title = Path(cell.title).name
        sheet_name = re.sub("[:]", "", title)
        view.df.to_excel(excel_writer, sheet_name=sheet_name, columns=view.columns, index=True)
        book = excel_writer.book
        sheet = excel_writer.sheets[sheet_name]

        cell_fill_df = self._gen_fill_styles(view)
        nindex = len(view.df.index.names)
        nheader = len(view.df.columns.names)

        side_thin = Side(border_style="dashed", color="000000")
        side_medium = Side(border_style="thin", color="000000")
        side_thick = Side(border_style="thick", color="000000")
        border_thin = Border(top=side_thin, bottom=side_thin, left=side_thin, right=side_thin)
        border_medium = Border(top=side_thin, bottom=side_thin, left=side_medium, right=side_thin)
        border_thick = Border(top=side_thin, bottom=side_thin, left=side_thick, right=side_thin)

        # Set column colors and width
        # Note that this will also adjust the column headers width
        for (idx, column), xcells in zip(enumerate(view.columns),
                                         sheet.iter_cols(min_row=nheader + 1, min_col=nindex + 1)):
            col_idx = nindex + idx + 1
            col_width = max(map(len, column))
            # We introduce a thicker border for column groups at the first two column index levels
            if len(view.columns.names) > 1 and idx > 0:
                if len(view.columns.names) > 2 and column[1] != view.columns[idx - 1][1]:
                    # generally the 'delta' or 'aggregate' levels
                    cell_border = border_medium
                elif column[0] != view.columns[idx - 1][0]:
                    # generally the 'metric' level
                    cell_border = border_thick
                else:
                    cell_border = border_thin
            else:
                cell_border = border_thin

            for df_index, xcell in zip(view.df.index, xcells):
                if pd.api.types.is_float_dtype(type(xcell.value)):
                    cell_width = len(f"{xcell.value:.2f}")
                    xcell.number_format = "0.00"
                else:
                    cell_width = len(str(xcell.value))
                if cell_fill_df is not None:
                    column_set = cell_fill_df.loc[[df_index], :]
                    if "column" in column_set:
                        fill = column_set.loc[column_set["column"] == column, "fill_style"]
                        assert len(fill) == 1
                    else:
                        fill = column_set["fill_style"]
                        assert len(fill) == 1
                    xcell.fill = fill.iloc[0]
                col_width = max(col_width, cell_width)
                xcell.border = cell_border
            sheet.column_dimensions[get_column_letter(col_idx)].width = col_width + 2

        # # Resize index columns to fit text
        for idx, level in enumerate(view.df.index.names):
            col_idx = idx + 1
            width = view.df.index.get_level_values(level).map(str).map(len).max()
            if level is None:
                level = ""
            width = max(width, len(level)) + 2  # some padding
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze index and headers, +1 is required because sheet columns start from 1, not zero.
        sheet.freeze_panes = sheet.cell(row=len(view.df.columns.names) + 1, column=len(view.df.index.names) + 1)
