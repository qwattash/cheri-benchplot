import re
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from matplotlib import colors as mcolors
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils import get_column_letter

from .plot import CellData, DataView, PlotError, PlotUnsupportedError, Surface


class SpreadsheetSurface(Surface):
    """
    Draw plots to static HTML files.
    """
    @dataclass
    class DrawContext(Surface.DrawContext):
        writer: pd.ExcelWriter

    def _make_draw_context(self, title, dest, **kwargs):
        writer = pd.ExcelWriter(dest.with_suffix(".xlsx"), mode="w", engine="openpyxl")
        return super()._make_draw_context(title, dest, writer=writer, **kwargs)

    def _finalize_draw_context(self, ctx):
        ctx.writer.close()

    def make_cell(self, **kwargs):
        return SpreadsheetPlotCell(**kwargs)

    def make_view(self, plot_type, **kwargs):
        if plot_type == "table":
            return SpreadsheetTable(**kwargs)
        raise PlotUnsupportedError(f"Plot type {plot_type} is not supported by the spreadsheet surface")


class SpreadsheetPlotCell(CellData):
    """
    Base HTML dataset rendering class. Add wrapper functions to allow running the rendering
    step within the jinja templates, so that we can access cell and view properties within the
    template if needed.
    """
    def draw(self, ctx: SpreadsheetSurface.DrawContext):
        name = self.title
        if len(self.views) > 1:
            self.logger.warning("Only a single plot view is supported for each excel surface cell")
        if len(self.views):
            self.views[0].render(self, self.surface, ctx.writer)


class SpreadsheetTable(DataView):
    """
    Render table in a spreadsheet.
    Note that this assumes that the sheet will only contain this table.
    """
    def _get_cell_color_styles(self, legend_map):
        fill_styles = {}
        if legend_map is None:
            return fill_styles

        for key, color in legend_map.color_items():
            # openpyxl uses aRGB colors while matplotlib uses RGBa
            # we need to swap the alpha here
            rgba_hex = [int(round(c * 255)) for c in color]
            argb_hex = rgba_hex[-1:] + rgba_hex[0:3]
            hexcolor = "".join(map(lambda c: f"{c:02x}", argb_hex))
            style = PatternFill(start_color=hexcolor, end_color=hexcolor, fill_type="solid")
            fill_styles[key] = style
        return fill_styles

    def render(self, cell, surface, excel_writer):
        """
        Render the dataframe as a table in an excel sheet.
        """
        if self.yright:
            surface.logger.warning("Excel table does not support right Y axis")
        if self.colormap:
            surface.logger.warning("Excel table does not support per-sample colormap")

        title = Path(cell.title).name
        sheet_name = re.sub("[:]", "", title)
        self.df.to_excel(excel_writer, sheet_name=sheet_name, columns=self.yleft, index=True)
        book = excel_writer.book
        sheet = excel_writer.sheets[sheet_name]

        fill_styles = self._get_cell_color_styles(cell.legend_map)
        nindex = len(self.df.index.names)
        nheader = len(self.df.columns.names)

        side_thin = Side(border_style="dashed", color="000000")
        side_medium = Side(border_style="thin", color="000000")
        side_thick = Side(border_style="thick", color="000000")
        border_thin = Border(top=side_thin, bottom=side_thin, left=side_thin, right=side_thin)
        border_medium = Border(top=side_thin, bottom=side_thin, left=side_medium, right=side_thin)
        border_thick = Border(top=side_thin, bottom=side_thin, left=side_thick, right=side_thin)

        # Set column colors and width
        # Note that this will also adjust the column headers width
        for (idx, column), xcells in zip(enumerate(self.yleft), sheet.iter_cols(min_row=nheader + 1,
                                                                                min_col=nindex + 1)):
            col_idx = nindex + idx + 1
            col_width = max(map(len, column))
            # XXX This should be a column_groups argument
            aggregate_level = self.df.columns.names.index("aggregate")
            metric_level = self.df.columns.names.index("metric")
            if idx > 0 and column[aggregate_level] != self.yleft[idx - 1][aggregate_level]:
                cell_border = border_medium
            elif idx > 0 and column[metric_level] != self.yleft[idx - 1][metric_level]:
                cell_border = border_thick
            else:
                cell_border = border_thin

            for xcell in xcells:
                if pd.api.types.is_float_dtype(type(xcell.value)):
                    cell_width = len(f"{xcell.value:.2f}")
                    xcell.number_format = "0.00"
                else:
                    cell_width = len(str(xcell.value))
                if cell.legend_map:
                    xcell.fill = fill_styles[column]
                col_width = max(col_width, cell_width)
                xcell.border = cell_border
            sheet.column_dimensions[get_column_letter(col_idx)].width = col_width + 2

        # # Resize index columns to fit text
        for idx, level in enumerate(self.df.index.names):
            col_idx = idx + 1
            width = self.df.index.get_level_values(level).map(str).map(len).max()
            if level is None:
                level = ""
            width = max(width, len(level)) + 2  # some padding
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze index and headers, +1 is required because sheet columns start from 1, not zero.
        sheet.freeze_panes = sheet.cell(row=len(self.df.columns.names) + 1, column=len(self.df.index.names) + 1)
